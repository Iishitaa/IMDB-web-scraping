import requests
from bs4 import BeautifulSoup 
import random
import openpyxl

exc = openpyxl.Workbook()
#print(exc.sheetnames)
sheet = exc.active
sheet.title = "Top rated IMDB movies"
#print(exc.sheetnames)

sheet.append(["Rank","Movie","Year","IMDB Rating"])


user_agents_list = [
    'Mozilla/5.0 (iPad; CPU OS 12_2 like Mac OS X) AppleWebKit/605.1.15 (KHTML, like Gecko) Mobile/15E148',
    'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/99.0.4844.83 Safari/537.36',
    'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/99.0.4844.51 Safari/537.36'
]


try:
    r = requests.get('https://www.imdb.com/chart/top/', headers={'User-Agent': random.choice(user_agents_list)})
    r.raise_for_status()
    soup = BeautifulSoup(r.text, "html.parser")
    #print(soup)


    movies = soup.find("tbody", class_= "lister-list").find_all("tr")
    for movie in movies:
        name = movie.find("td", class_="titleColumn").a.text
        rank =  movie.find("td", class_="titleColumn").get_text(strip = True).split(".")[0]
        year = movie.find("td", class_="titleColumn").span.text.strip("()")
        rating = movie.find("td", class_="ratingColumn imdbRating").strong.text
        print(rank, name, year, rating)

        sheet.append([rank,name,year,rating])


except Exception as e:
    print(e)
    
exc.save("IMDB web scarping.xlsx")



